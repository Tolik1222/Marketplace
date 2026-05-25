from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django import forms
from .forms import UserRegistrationForm, UserLoginForm, UserProfileForm
from products.models import Product
from orders.models import Order


def register(request):
    """Реєстрація нового користувача"""
    if request.user.is_authenticated:
        return redirect('accounts:profile')
    
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Вітаємо! Ви успішно зареєструвались.')
            return redirect('accounts:profile')
        else:
            messages.error(request, 'Будь ласка, виправте помилки в формі.')
    else:
        form = UserRegistrationForm()
    
    return render(request, 'accounts/register.html', {'form': form})


def user_login(request):
    """Вхід користувача"""
    if request.user.is_authenticated:
        return redirect('accounts:profile')
    
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            
            if user is not None:
                login(request, user)
                messages.success(request, f'Вітаємо, {user.username}!')
                
                # кидаємо туди, звідки прийшов, або в профіль
                next_page = request.GET.get('next', 'accounts:profile')
                return redirect(next_page)
        else:
            messages.error(request, 'Невірне ім\'я користувача або пароль.')
    else:
        form = UserLoginForm()
    
    return render(request, 'accounts/login.html', {'form': form})


@login_required
def user_logout(request):
    """Вихід користувача"""
    logout(request)
    messages.info(request, 'Ви успішно вийшли з системи.')
    return redirect('products:product_list')


@login_required
def profile(request):
    """Особистий кабінет користувача"""
    user_products = Product.objects.filter(owner=request.user) if hasattr(Product, "owner") else Product.objects.none()
    user_orders = (
        Order.objects.filter(Q(user=request.user) | Q(user__isnull=True, email=request.user.email))
        .prefetch_related("items__product")
        .order_by("-created")
        .distinct()
    )

    context = {
        "user_products": user_products,
        "user_orders": user_orders,
    }
    return render(request, 'accounts/profile.html', context)


@login_required
def profile_edit(request):
    """Редагування профілю"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Профіль успішно оновлено!')
            return redirect('accounts:profile')
        else:
            messages.error(request, 'Будь ласка, виправте помилки.')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'accounts/profile_edit.html', {'form': form})


from decimal import Decimal
from django.core.exceptions import PermissionDenied
from orders.models import OrderItem

@login_required
def seller_dashboard(request):
    if not request.user.is_staff:
        raise PermissionDenied("Ви повинні бути продавцем, щоб переглядати цю сторінку.")

    products = Product.objects.filter(owner=request.user).order_by("-updated")
    total_products = products.count()

    # Optimize N+1 queries by prefetching related order items and selecting coupons
    order_items = OrderItem.objects.filter(
        product__owner=request.user
    ).select_related('order', 'order__coupon', 'product').prefetch_related('order__items')
    
    total_sold = 0
    total_revenue = Decimal("0.00")
    orders_dict = {}

    for item in order_items:
        if item.order.paid:
            total_sold += item.quantity
            
            # рахуємо чистий дохід з урахуванням знижки
            total_cost = item.order.get_total_cost()
            if total_cost > 0:
                discount_ratio = item.order.get_discount_amount() / total_cost
                discount_multiplier = Decimal("1") - discount_ratio
            else:
                discount_multiplier = Decimal("1")
                
            item_revenue = (item.price * item.quantity * discount_multiplier).quantize(Decimal("0.01"))
            total_revenue += item_revenue

        order = item.order
        if order.id not in orders_dict:
            orders_dict[order.id] = {
                'order': order,
                'items': [],
                'seller_total': Decimal("0.00"),
            }
        
        # рахуємо суму для поточного замовлення
        total_cost = order.get_total_cost()
        if total_cost > 0:
            discount_ratio = order.get_discount_amount() / total_cost
            discount_multiplier = Decimal("1") - discount_ratio
        else:
            discount_multiplier = Decimal("1")
            
        item_revenue = (item.price * item.quantity * discount_multiplier).quantize(Decimal("0.01"))
        orders_dict[order.id]['items'].append(item)
        orders_dict[order.id]['seller_total'] += item_revenue

    seller_orders = list(orders_dict.values())
    seller_orders.sort(key=lambda x: x['order'].created, reverse=True)

    from orders.models import SubOrder
    seller_sub_orders = SubOrder.objects.filter(
        vendor=request.user
    ).select_related('order', 'order__coupon').prefetch_related('items__product').order_by('-created')

    context = {
        'products': products,
        'total_products': total_products,
        'total_sold': total_sold,
        'total_revenue': total_revenue,
        'seller_orders': seller_orders,
        'seller_sub_orders': seller_sub_orders,
    }
    return render(request, 'accounts/seller_dashboard.html', context)


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import now

@login_required
def seller_export_excel(request):
    if not request.user.is_staff:
        raise PermissionDenied("Ви повинні бути продавцем, щоб виконувати цю дію.")

    # створюємо нову книгу і лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Звіт про продажі"
    
    # показуємо сітку
    ws.views.sheetView[0].showGridLines = True

    # заголовки колонок
    headers = [
        "ID Замовлення", 
        "Дата", 
        "Покупець", 
        "Служба доставки", 
        "Товар", 
        "Кількість", 
        "Ціна за од. (грн)", 
        "Дохід продавця (грн)", 
        "Статус"
    ]
    
    # стилі та кольори для таблиці
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10)
    font_total = Font(name="Segoe UI", size=11, bold=True)
    
    fill_header = PatternFill(fill_type="solid", start_color="1B3FAE", end_color="1B3FAE") # Brand accent navy
    fill_total = PatternFill(fill_type="solid", start_color="EDF2FF", end_color="EDF2FF") # Light soft blue
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )
    
    # записуємо заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # вибираємо продані товари цього продавця (з оптимізацією N+1)
    order_items = OrderItem.objects.filter(
        product__owner=request.user
    ).select_related('order', 'order__coupon', 'product').prefetch_related('order__items').order_by('-order__created')
    
    current_row = 2
    for item in order_items:
        order = item.order
        
        # рахуємо дохід з купонами
        total_cost = order.get_total_cost()
        if total_cost > 0:
            discount_ratio = order.get_discount_amount() / total_cost
            discount_multiplier = Decimal("1") - discount_ratio
        else:
            discount_multiplier = Decimal("1")
            
        item_revenue = float((item.price * item.quantity * discount_multiplier).quantize(Decimal("0.01")))
        
        # дані рядка
        row_values = [
            f"#{order.id}",
            order.created.strftime("%Y-%m-%d %H:%M"),
            f"{order.first_name} {order.last_name}",
            order.get_delivery_service_display() if hasattr(order, 'get_delivery_service_display') else "Не вказано",
            item.product.name,
            item.quantity,
            float(item.price),
            item_revenue,
            order.get_status_display() if hasattr(order, 'get_status_display') else ("Оплачено" if order.paid else "Очікує оплати")
        ]
        
        # пишемо в клітинки
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            
            # формати та вирівнювання
            if col_idx in [1, 2, 9]:
                cell.alignment = align_center
            elif col_idx in [3, 4, 5]:
                cell.alignment = align_left
            elif col_idx in [6]:
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx in [7, 8]:
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
                
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
    # рядок «разом»
    if current_row > 2:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        total_label_cell = ws.cell(row=current_row, column=1, value="ЗАГАЛЬНИЙ ДОХІД ПРОДАВЦЯ:")
        total_label_cell.font = font_total
        total_label_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_label_cell.fill = fill_total
        total_label_cell.border = thin_border
        
        for col_idx in range(2, 8):
            ws.cell(row=current_row, column=col_idx).fill = fill_total
            ws.cell(row=current_row, column=col_idx).border = thin_border
            
        sum_formula = f"=SUM(H2:H{current_row-1})"
        total_sum_cell = ws.cell(row=current_row, column=8, value=sum_formula)
        total_sum_cell.font = font_total
        total_sum_cell.alignment = align_right
        total_sum_cell.fill = fill_total
        total_sum_cell.border = thin_border
        total_sum_cell.number_format = '#,##0.00'
        
        ws.cell(row=current_row, column=9).fill = fill_total
        ws.cell(row=current_row, column=9).border = thin_border
        ws.row_dimensions[current_row].height = 24
    
    # підганяємо ширину колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "123,456.78"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # відправляємо файл користувачу
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="sales_report_{now().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response


import json
import uuid
import logging
import requests as http_requests
from django.core.exceptions import PermissionDenied
from django.conf import settings
from orders.models import SubOrder

logger_v = logging.getLogger(__name__)


@login_required
def seller_suborder_detail(request, sub_order_id):
    """Деталі підзамовлення для продавця + зміна статусу."""
    if not request.user.is_staff:
        raise PermissionDenied("Ви повинні бути продавцем.")

    sub_order = get_object_or_404(SubOrder, id=sub_order_id, vendor=request.user)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        allowed = ['paid', 'shipped', 'delivered', 'canceled']
        if new_status in allowed:
            sub_order.status = new_status
            sub_order.save()
            messages.success(request, f"Статус підзамовлення #{sub_order.id} оновлено на '{new_status}'.")
        return redirect('accounts:seller_suborder_detail', sub_order_id=sub_order.id)

    return render(request, 'accounts/suborder_detail.html', {'sub_order': sub_order})


@login_required
def generate_novaposhta_waybill(request, sub_order_id):
    """Автогенерація ТТН Нової Пошти для підзамовлення."""
    if not request.user.is_staff:
        raise PermissionDenied("Ви повинні бути продавцем.")

    sub_order = get_object_or_404(SubOrder, id=sub_order_id, vendor=request.user)
    profile = getattr(request.user, 'seller_profile', None)

    NP_API_KEY = getattr(settings, 'NOVA_POSHTA_API_KEY', None)
    generated_ttn = None
    error_msg = None

    if NP_API_KEY and profile and profile.np_sender_ref:
        order = sub_order.order
        payload = {
            "apiKey": NP_API_KEY,
            "modelName": "InternetDocument",
            "calledMethod": "save",
            "methodProperties": {
                "PayerType": "Sender",
                "PaymentMethod": "Cash",
                "DateTime": order.created.strftime("%d.%m.%Y"),
                "CargoType": "Cargo",
                "Weight": "1",
                "ServiceType": "WarehouseWarehouse",
                "SeatsAmount": "1",
                "Description": f"Замовлення #{order.id}",
                "Cost": str(int(sub_order.get_total_after_discount())),
                "CitySender": profile.np_sender_ref,
                "Sender": profile.np_sender_ref,
                "SenderAddress": profile.np_sender_address_ref or "",
                "ContactSender": profile.np_sender_contact_ref or "",
                "SendersPhone": profile.np_sender_phone or "",
                "CityRecipient": order.city,
                "RecipientAddress": order.branch,
                "RecipientsPhone": "",
                "RecipientName": f"{order.first_name} {order.last_name}",
            }
        }
        try:
            resp = http_requests.post("https://api.novaposhta.ua/v2.0/json/", json=payload, timeout=10)
            data = resp.json()
            if data.get("success") and data.get("data"):
                generated_ttn = data["data"][0].get("IntDocNumber", "")
                waybill_ref = data["data"][0].get("Ref", "")
                sub_order.tracking_number = generated_ttn
                sub_order.waybill_ref = waybill_ref
                sub_order.save()
                messages.success(request, f"ТТН успішно згенеровано: {generated_ttn}")
            else:
                errors = data.get("errors", [])
                error_msg = "; ".join(errors) if errors else "Невідома помилка API."
                logger_v.warning("Nova Poshta waybill generation failed for SubOrder %s: %s", sub_order_id, error_msg)
        except Exception as exc:
            error_msg = str(exc)
            logger_v.exception("Nova Poshta API error for SubOrder %s", sub_order_id)
    else:
        # Fallback: generate a test tracking number
        generated_ttn = f"59{str(uuid.uuid4().int)[:12]}"
        sub_order.tracking_number = generated_ttn
        sub_order.waybill_ref = "test"
        sub_order.save()
        messages.warning(request, f"API Нової Пошти не налаштовано. Тестовий трек-номер: {generated_ttn}")

    if error_msg:
        messages.error(request, f"Помилка генерації ТТН: {error_msg}")

    return redirect('accounts:seller_suborder_detail', sub_order_id=sub_order.id)


from .models import SellerProfile

class SellerProfileForm(forms.ModelForm):
    class Meta:
        model = SellerProfile
        fields = ['stripe_account_id', 'np_sender_ref', 'np_sender_address_ref',
                  'np_sender_contact_ref', 'np_sender_phone']
        labels = {
            'stripe_account_id': 'Stripe Connect Account ID',
            'np_sender_ref': 'Нова Пошта: Ref відправника',
            'np_sender_address_ref': 'Нова Пошта: Ref адреси відправника',
            'np_sender_contact_ref': 'Нова Пошта: Ref контакту відправника',
            'np_sender_phone': 'Нова Пошта: Телефон відправника',
        }
        widgets = {
            'stripe_account_id': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'acct_...'}),
            'np_sender_ref': forms.TextInput(attrs={'class': 'form-control'}),
            'np_sender_address_ref': forms.TextInput(attrs={'class': 'form-control'}),
            'np_sender_contact_ref': forms.TextInput(attrs={'class': 'form-control'}),
            'np_sender_phone': forms.TextInput(attrs={'class': 'form-control', 'placeholder': '+380...'}),
        }


@login_required
def seller_profile_edit(request):
    """Редагування налаштувань профілю продавця: Stripe Connect та Нова Пошта."""
    if not request.user.is_staff:
        raise PermissionDenied("Ви повинні бути продавцем.")

    profile, _ = SellerProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        form = SellerProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Профіль продавця успішно оновлено.")
            return redirect('accounts:seller_dashboard')
    else:
        form = SellerProfileForm(instance=profile)

    return render(request, 'accounts/seller_profile_edit.html', {'form': form})